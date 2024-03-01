import datetime
import os
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

Service_Obj = Service()
webdriver.Chrome(service=Service_Obj)

url = "https://reports.parkzap.com/login"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get(url)
print(driver.title)
print(driver.current_url)
driver.maximize_window()

# time.sleep(1)

# Sending the User Email ID
time.sleep(2)
driver.find_element(By.XPATH, "//input[@placeholder='Email']").send_keys("abhishek.sahu@stackfusion.io")
print("Sending User ID")

driver.find_element(By.XPATH, "//input[@placeholder='Password']").send_keys("Harrypotter@2023")
print("Sending User Pass")

driver.find_element(By.XPATH, "//button[normalize-space()='Sign in']").click()
print("Clicking on Sign IN Button")
# time.sleep(2)

driver.find_element(By.XPATH, "(//span[@class='p-dropdown-trigger-icon p-clickable pi pi-chevron-down'])[1]").click()
print("Clicking on Dropdown")
# time.sleep(2)

# driver.find_element(By.CLASS_NAME, "p-dropdown-filter").click()
driver.find_element(By.CLASS_NAME, "p-dropdown-filter").send_keys("Ahm")
print("Clicking on dropdown and sending the input")

# time.sleep(2)
# It's selecting the site which we have selected
sites = driver.find_elements(By.CSS_SELECTOR, "li[class='p-dropdown-item']")
print(len(sites))
# time.sleep(2)

assert len(sites) > 1, "No sites found in the dropdown"

found_site = False
for site_list in sites:
    if site_list.text == "Ahmedabad Airport T2":
        found_site = True
        print(found_site)
        site_list.click()
        time.sleep(2)
        break

# Assert that "Ahmedabad Airport T2" was found
assert found_site, "Site 'Ahmedabad Airport T2' not found in the dropdown"
print("Site 'Ahmedabad Airport T2' selected successfully")

driver.find_element(By.XPATH, "//span[normalize-space()='Day Close Report']").click()
driver.find_element(By.CLASS_NAME, "p-button").click()
print("Clicked on Calender")

# SELECTING THE GIVEN DATE

driver.find_element(By.XPATH, "//span[normalize-space()='6']").click()

time.sleep(4)

# TO DELETE EXISTING FILE FROM OS

DOWNLOAD_PATH = "/home/abhishek/Downloads"

FILE_NAME = "Day Close Report.csv"

file_path = os.path.join(DOWNLOAD_PATH, FILE_NAME)
if os.path.exists(file_path):
    os.remove(file_path)
    print(f"Removed existing file: {FILE_NAME}")

get_report = driver.find_elements(By.CLASS_NAME, "p-button-label")
print(len(get_report))

assert len(get_report) > 1, "No module found"
found_getReport = True
for get_report in get_report:
    if get_report.text == "Get Report":
        # get_report = True
        print(found_getReport)
        get_report.click()
        break

assert found_getReport, "Site 'Get Report' not found in the dropdown"
print("Site 'Ahmedabad Airport T2' selected successfully")

# driver.find_element(By.CSS_SELECTOR, "#pr_id_20").click()

time.sleep(3)

# REPORTS DOWNLOAD PATH

driver.find_element(By.XPATH, "//span[normalize-space()='Export CSV']").click()
print("*************Reports Generated*************")

# NOW ACCESSING THE TICKET REPORT TAB
driver.find_element(By.XPATH, "//a[@data-pr-tooltip='Ticket Report']").click()

time.sleep(3)

# CLICKED ON CALENDER
driver.find_element(By.CLASS_NAME, "p-button-icon-only").click()

# SELECTED THE ABSOLUTE DATE
Date_Fields = driver.find_elements(By.CLASS_NAME, "p-listbox-item")
print(len(Date_Fields))
assert len(Date_Fields) > 1, "No Field Available"
Found_Date_Field = False

for Dates in Date_Fields:
    if Dates.text == "Absolute date":
        Found_Date_Field = True
        print(Found_Date_Field)
        Dates.click()
        break

assert Found_Date_Field, "Absolute date field not found in the list"

time.sleep(3)

# # NOW SELECTING THE DATE IN TICKET REPORT Dynamically
# current_date = datetime.date.today()
#
# date_format = "%-m/%-d/%Y"
#
# formatted_date = current_date.strftime(date_format)
#
# driver.find_element(By.CLASS_NAME, "p-datepicker-today").click()
# time.sleep(5)

action = ActionChains(driver)

action.click_and_hold(driver.find_element(By.XPATH, "//span[@class='p-highlight']")).perform()
driver.find_element(By.XPATH, "//span[normalize-space()='6']").click()
driver.find_element(By.XPATH, "//span[normalize-space()='7']").click()
time.sleep(4)

driver.find_element(By.XPATH, "//button[normalize-space()='Apply']").click()

# This will remove the Previous Report
DOWNLOAD_PATH_Ticket_Report = "/home/abhishek/Downloads"

FILE_NAME = "Detailed Ticket Log.csv"

file_path = os.path.join(DOWNLOAD_PATH, FILE_NAME)
if os.path.exists(file_path):
    os.remove(file_path)
    print(f"Removed existing file: {FILE_NAME}")

time.sleep(1)
Ticket_Report = driver.find_element(By.XPATH, "//span[normalize-space()='Get Report']")
Ticket_Report.click()
print(Ticket_Report)
time.sleep(10)
driver.find_element(By.XPATH, "//button[normalize-space()='Export CSV']").click()
time.sleep(2)
print("Detailed Reported Generated")

time.sleep(5)

exportDayCloseReport = openpyxl.load_workbook("/home/abhishek/Downloads/Day Close Report.xlsx")
sheet = exportDayCloseReport.active

# cell = sheet.cell(row=1, column=5)
# print(cell.value)
# print(sheet.max_row)
# print(sheet.max_column)
#
# # for i in range(1,sheet.max_row+1):
# #     for j in range(1,sheet.max_column+1):
# #         print(sheet.cell(row=i, column=j).value)
# for i in range(1, sheet.max_row + 1):
#     for j in range(1, sheet.max_column + 1):
#         cell_value = sheet.cell(row=i, column=j).value
#         if cell_value == "Total":
#             print(f"Found 'Total' at row {i}, column {j}")
#         else:
#             print(cell_value)
# cell = sheet.cell(row=1, column=5)
# print(f"Header: {cell.value}")
# print(f"Max Rows: {sheet.max_row}")
# print(f"Max Columns: {sheet.max_column}")
#
# for i in range(1, sheet.max_row + 1):
#     row_values = [str(sheet.cell(row=i, column=j).value) for j in range(1, sheet.max_column + 1)]
#
#     # Print each row on a new line with values separated by tabs
#     print("\t".join(row_values))

# ... (previous code)

# Assuming 'E' is the column you want to sum
target_column = 'E'

# Find the index of the target column in the header row
target_column_index = None
for j in range(1, sheet.max_column + 1):
    if sheet.cell(row=1, column=j).column_letter == target_column:
        target_column_index = j
        break

# Check if the target column was found
if target_column_index is not None:
    # Initialize sum
    total_sum = 0

    # Iterate through the rows and accumulate values in the target column
    for i in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=i, column=target_column_index).value
        # Ensure the cell contains a numeric value before adding
        if isinstance(cell_value, (int, float)):
            total_sum += cell_value

    # Print the total sum
    print(f"Total sum in column '{target_column}': {total_sum:.2f}")
else:
    print(f"Column '{target_column}' not found.")

exportTicketReport = openpyxl.load_workbook("/home/abhishek/Downloads/Detailed Ticket Log.xlsx")
sheet = exportTicketReport.active

# cell = sheet.cell(row=1, column=5)
# print(cell.value)
# print(sheet.max_row)
# print(sheet.max_column)
#
# for i in range(1, sheet.max_row + 1):
#     for j in range(1, sheet.max_column + 1):
#         cell_value = sheet.cell(row=i, column=j).value
#         if cell_value == "Total":
#             print(f"Found 'Total' at row {i}, column {j}")
#         else:
#             print(cell_value)
# cell = sheet.cell(row=1, column=5)
# print(f"Header: {cell.value}")
# print(f"Max Rows: {sheet.max_row}")
# print(f"Max Columns: {sheet.max_column}")
#
# for i in range(1, sheet.max_row + 1):
#     row_values = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
#
#     # Print each row on a new line
#     print(" ".join(map(str, row_values)))

# Find the index of the "Amount Paid Electronically" column
amount_paid_index = None
for j in range(1, sheet.max_column + 1):
    if sheet.cell(row=1, column=j).value == "Amount Paid Electronically":
        amount_paid_index = j
        break

# Check if the column was found
if amount_paid_index is not None:
    # Initialize sum
    total_amount_paid = 0

    # Iterate through the rows and accumulate values in the column
    for i in range(2, sheet.max_row + 1):  # Assuming data starts from row 2
        cell_value = sheet.cell(row=i, column=amount_paid_index).value
        # Ensure the cell contains a numeric value before adding
        if isinstance(cell_value, (int, float)):
            total_amount_paid += cell_value

    # Print the total amount paid
    print(f"Total Amount Paid Electronically: ₹{total_amount_paid:.2f}")
else:
    print("Column 'Amount Paid Electronically' not found.")
